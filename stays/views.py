import json
from urllib.parse import urlencode

from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Count, Sum, Avg, Q
from django.db.models.functions import TruncDate
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.utils.encoding import smart_str
import csv
from datetime import datetime
from pathlib import Path
from django.conf import settings
from django.contrib import messages

from .models import Stay
from decimal import Decimal
from datetime import datetime
from django.db import models as dj_models

# Try to use your app's form; fallback to a simple ModelForm
try:
    from .forms import StayForm
except Exception:
    from django.forms import ModelForm
    class StayForm(ModelForm):
        class Meta:
            model = Stay
            fields = "__all__"

def _apply_stay_filters(qs, request):
    """Apply multi-filters: state, city, rating (if rating field exists)."""
    states = request.GET.getlist("state") or ([request.GET.get("state")] if request.GET.get("state") else [])
    cities = request.GET.getlist("city") or ([request.GET.get("city")] if request.GET.get("city") else [])
    ratings = request.GET.getlist("rating") or ([request.GET.get("rating")] if request.GET.get("rating") else [])
    q = (request.GET.get("q") or "").strip()
    start = (request.GET.get("start") or "").strip()
    end = (request.GET.get("end") or "").strip()
    paid = (request.GET.get("paid") or "").strip()
    min_price = (request.GET.get("min_price") or "").strip()
    max_price = (request.GET.get("max_price") or "").strip()
    missing_coords = (request.GET.get("missing_coords") or "").strip()
    years = request.GET.getlist("year") or ([request.GET.get("year")] if request.GET.get("year") else [])

    states = [s for s in states if s]
    cities = [c for c in cities if c]
    ratings_clean = []
    for r in ratings:
        try:
            ratings_clean.append(int(r))
        except Exception:
            pass

    if states:
        states_up = [s.upper() for s in states]
        qs = qs.filter(state__in=states_up)
    if cities:
        qs = qs.filter(city__in=cities)

    field_names = {getattr(f, "attname", None) or getattr(f, "name", None) for f in Stay._meta.get_fields()}
    if ratings_clean and "rating" in field_names:
        qs = qs.filter(rating__in=ratings_clean)

    if q:
        qs = qs.filter(
            Q(park__icontains=q)
            | Q(city__icontains=q)
            | Q(state__icontains=q)
            | Q(site__icontains=q)
            | Q(notes__icontains=q)
        )

    # Dates: filter check_in >= start, leave_date <= end
    from datetime import date
    def _parse_date(s: str):
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        return None
    d_start = _parse_date(start)
    d_end = _parse_date(end)
    if d_start:
        qs = qs.filter(check_in__isnull=False, check_in__gte=d_start)
    if d_end:
        qs = qs.filter(leave_date__isnull=False, leave_date__lte=d_end)

    if paid in {"0", "1"}:
        qs = qs.filter(paid=(paid == "1"))
    # Filter by year(s) across check_in or leave_date
    years_int = []
    for y in years:
        try:
            years_int.append(int(str(y).strip()))
        except Exception:
            pass
    if years_int:
        qy = Q()
        qy |= Q(check_in__year__in=years_int)
        qy |= Q(leave_date__year__in=years_int)
        qs = qs.filter(qy)

    from decimal import Decimal, InvalidOperation
    def _dec(s):
        if not s:
            return None
        try:
            return Decimal(s)
        except InvalidOperation:
            return None
    dmin = _dec(min_price)
    dmax = _dec(max_price)
    if dmin is not None:
        qs = qs.filter(price_night__isnull=False, price_night__gte=dmin)
    if dmax is not None:
        qs = qs.filter(price_night__isnull=False, price_night__lte=dmax)

    # Missing coordinates
    if missing_coords in {"1", "true", "yes"}:
        qs = qs.filter(Q(latitude__isnull=True) | Q(longitude__isnull=True))

    return qs

def stay_list(request):
    qs = Stay.objects.all()
    # Read selections first so we can drive dependent choices
    selected_states  = request.GET.getlist("state")  or ([request.GET.get("state")]  if request.GET.get("state")  else [])
    selected_cities  = request.GET.getlist("city")   or ([request.GET.get("city")]   if request.GET.get("city")   else [])
    selected_ratings = request.GET.getlist("rating") or ([request.GET.get("rating")] if request.GET.get("rating") else [])
    selected_q = (request.GET.get("q") or "").strip()
    selected_start = (request.GET.get("start") or "").strip()
    selected_end = (request.GET.get("end") or "").strip()
    selected_paid = (request.GET.get("paid") or "").strip()
    selected_min_price = (request.GET.get("min_price") or "").strip()
    selected_max_price = (request.GET.get("max_price") or "").strip()
    selected_sort = (request.GET.get("sort") or "date_desc").strip()
    selected_missing = (request.GET.get("missing_coords") or "").strip() in {"1","true","yes"}
    selected_years = request.GET.getlist("year") or ([request.GET.get("year")] if request.GET.get("year") else [])
    selected_ratings = [str(r) for r in selected_ratings]

    # Choices
    state_choices = list(Stay.objects.values_list("state", flat=True)
                         .exclude(state__isnull=True).exclude(state__exact="")
                         .distinct().order_by("state"))
    if selected_states:
        states_up = [s.upper() for s in selected_states]
        city_base = Stay.objects.filter(state__in=states_up)
    else:
        city_base = Stay.objects
    city_choices  = list(city_base.values_list("city", flat=True)
                         .exclude(city__isnull=True).exclude(city__exact="")
                         .distinct().order_by("city"))
    rating_choices = [1, 2, 3, 4, 5]
    # Year choices
    year_set = set()
    try:
        for d in Stay.objects.exclude(check_in__isnull=True).dates("check_in", "year"):
            year_set.add(d.year)
    except Exception:
        pass
    try:
        for d in Stay.objects.exclude(leave_date__isnull=True).dates("leave_date", "year"):
            year_set.add(d.year)
    except Exception:
        pass
    year_choices = sorted(year_set, reverse=True)

    # Apply filters to listing queryset
    qs = _apply_stay_filters(qs, request)

    # Sorting
    sort_map = {
        "date_desc": ("-check_in", "-id"),
        "date_asc": ("check_in", "id"),
        "price_desc": ("-price_night", "-id"),
        "price_asc": ("price_night", "id"),
        "rating_desc": ("-rating", "-id"),
        "rating_asc": ("rating", "id"),
        "park_asc": ("park", "id"),
    }
    order = sort_map.get(selected_sort) or sort_map["date_desc"]
    qs = qs.order_by(*order)

    # Build mapping of state -> cities for client-side dynamic filtering
    pairs = (Stay.objects
             .values_list("state", "city")
             .exclude(state__isnull=True).exclude(state__exact="")
             .exclude(city__isnull=True).exclude(city__exact="")
             .distinct())
    cities_by_state = {}
    for st, ct in pairs:
        st_up = (st or "").upper()
        cities_by_state.setdefault(st_up, set()).add(ct)
    cities_by_state = {k: sorted(v) for k, v in cities_by_state.items()}

    qs_params = []
    for s in selected_states:  qs_params.append(("state", s))
    for c in selected_cities:  qs_params.append(("city", c))
    for r in selected_ratings: qs_params.append(("rating", r))
    for y in selected_years: qs_params.append(("year", y))
    if selected_q: qs_params.append(("q", selected_q))
    if selected_start: qs_params.append(("start", selected_start))
    if selected_end: qs_params.append(("end", selected_end))
    if selected_paid in {"0","1"}: qs_params.append(("paid", selected_paid))
    if selected_min_price: qs_params.append(("min_price", selected_min_price))
    if selected_max_price: qs_params.append(("max_price", selected_max_price))
    if selected_sort: qs_params.append(("sort", selected_sort))
    if selected_missing: qs_params.append(("missing_coords", "1"))
    map_query = urlencode(qs_params)

    return render(request, "stays/stay_list.html", {
        "stays": qs,
        "state_choices": state_choices,
        "city_choices": city_choices,
        "cities_by_state": json.dumps(cities_by_state),
        "rating_choices": rating_choices,
        "selected_states": selected_states,
        "selected_cities": selected_cities,
        "selected_ratings": selected_ratings,
        "selected_years": [str(y) for y in selected_years],
        "selected_q": selected_q,
        "selected_start": selected_start,
        "selected_end": selected_end,
        "selected_paid": selected_paid,
        "selected_min_price": selected_min_price,
        "selected_max_price": selected_max_price,
        "selected_sort": selected_sort,
        "selected_missing": selected_missing,
        "year_choices": year_choices,
        "map_query": map_query,
        "stars": [1, 2, 3, 4, 5],
    })

def stay_map(request):
    """Leaflet map of stays; zooms to filtered stays if filters present, otherwise all."""
    qs = _apply_stay_filters(Stay.objects.all(), request)
    qs = qs.exclude(latitude__isnull=True).exclude(longitude__isnull=True)
    points = []
    for s in qs:
        lat = float(s.latitude) if s.latitude is not None else None
        lng = float(s.longitude) if s.longitude is not None else None
        title = f"{getattr(s, 'park', '')}".strip() or f"Stay {s.pk}"
        city = f"{getattr(s, 'city', '')}".strip()
        state = f"{getattr(s, 'state', '')}".strip()
        subtitle = f"{city}, {state}".strip(", ")
        points.append({
            # Our newer template keys
            "lat": lat,
            "lng": lng,
            "title": title,
            "subtitle": subtitle,
            "pk": s.pk,
            # Legacy project-level template keys
            "latitude": lat,
            "longitude": lng,
            "name": title,
            "city": city,
            "state": state,
        })
    return render(request, "stays/map.html", {"stays_json": json.dumps(points)})


def appearance_page(request):
    # Clear any stale success/info messages from prior actions so this page stays clean
    try:
        from django.contrib import messages as _msgs
        list(_msgs.get_messages(request))
    except Exception:
        pass
    return render(request, "stays/appearance.html")

def appearance_geocode(request):
    """Bulk geocode from Appearance page; stays on the same page when done."""
    if request.method != 'POST':
        return HttpResponseBadRequest('POST required')
    try:
        limit_param = request.POST.get('limit') or request.GET.get('limit') or '1000'
        limit = None if str(limit_param).lower() in {'all', 'unlimited', 'none'} else int(limit_param)
    except Exception:
        limit = 1000
    from .utils import geocode_from_stay
    qs = Stay.objects.filter(latitude__isnull=True) | Stay.objects.filter(longitude__isnull=True)
    updated = 0
    # Apply limit if provided
    iterable = qs.iterator() if limit is None else qs[:limit]
    for s in iterable:
        coords = geocode_from_stay(s)
        if coords:
            s.latitude, s.longitude = coords
            s.save(update_fields=['latitude', 'longitude'])
            updated += 1
    messages.success(request, f"Geocoded {updated} stay(s).")
    return redirect('stays:appearance')

def appearance_purge(request):
    """Delete all Stay rows (POST only). Use with care."""
    if request.method != 'POST':
        return HttpResponseBadRequest('POST required')
    total = Stay.objects.count()
    Stay.objects.all().delete()
    messages.success(request, f"Purged {total} stay(s).")
    return redirect('stays:appearance')

def appearance_normalize_cities(request):
    """Normalize city values by removing trailing ", ST" and filling state if missing."""
    if request.method != 'POST':
        return HttpResponseBadRequest('POST required')
    import re
    fixed = 0
    filled_state = 0
    for s in Stay.objects.all():
        city = (s.city or '').strip()
        if not city:
            continue
        m = re.search(r"^(.*?),(?:\s*)([A-Za-z]{2})$", city)
        changed = False
        if m:
            base, suff = m.group(1).strip(), m.group(2).upper()
            if not (s.state or '').strip():
                s.state = suff
                filled_state += 1
                changed = True
            if base != s.city:
                s.city = base
                changed = True
        elif "," in city:
            base = city.split(",", 1)[0].strip()
            if base != s.city:
                s.city = base
                changed = True
        else:
            # Case: "City ST" without comma
            parts = city.split()
            if len(parts) >= 2 and len(parts[-1]) == 2:
                base = " ".join(parts[:-1]).strip()
                if not (s.state or '').strip():
                    s.state = parts[-1].upper()
                    filled_state += 1
                    changed = True
                if base != s.city:
                    s.city = base
                    changed = True
        if changed:
            s.save(update_fields=['city', 'state'])
            fixed += 1
    messages.success(request, f"Normalized {fixed} stay(s); filled state for {filled_state}.")
    return redirect('stays:appearance')

def geocode_missing(request):
    """Geocode stays missing coordinates. POST only. Optional ?limit=..."""
    if request.method != 'POST':
        return HttpResponseBadRequest('POST required')
    try:
        limit = int(request.GET.get('limit', '100'))
    except Exception:
        limit = 100
    from .utils import build_query_from_stay, geocode_address
    qs = Stay.objects.filter(latitude__isnull=True) | Stay.objects.filter(longitude__isnull=True)
    updated = 0
    for s in qs[:limit]:
        q = build_query_from_stay(s)
        if not q:
            continue
        coords = geocode_address(q)
        if coords:
            s.latitude, s.longitude = coords
            s.save(update_fields=['latitude', 'longitude'])
            updated += 1
    messages.success(request, f"Geocoded {updated} stay(s).")
    return redirect('stays:map')

def stay_detail(request, pk):
    obj = get_object_or_404(Stay, pk=pk)
    # Build a list of (label, value) for all concrete model fields
    all_fields = []
    for f in Stay._meta.fields:
        # Skip photo here; it's already shown at the top
        if f.name == "photo":
            continue
        label = getattr(f, "verbose_name", f.name) or f.name
        try:
            label = str(label).capitalize()
        except Exception:
            label = f.name
        val = getattr(obj, f.name, None)
        # Pretty formatting by field type/name
        display = None
        try:
            # Dates
            if isinstance(f, dj_models.DateField) and val:
                display = val.strftime("%Y-%m-%d")
            # Decimals (money/coords)
            elif isinstance(f, dj_models.DecimalField) and val is not None:
                if f.name in {"latitude", "longitude"}:
                    display = f"{float(val):.6f}"
                elif f.name in {"price_night", "total", "fees"}:
                    display = f"{float(val):.2f}"
            # Booleans
            elif isinstance(f, dj_models.BooleanField):
                display = "Yes" if bool(val) else "No"
        except Exception:
            display = None
        if display is None:
            display = val if (val not in (None, "")) else "—"
        all_fields.append((label, display))
    # Include a useful derived value
    try:
        all_fields.append(("Nights", obj.nights))
    except Exception:
        pass

    return render(
        request,
        "stays/stay_detail.html",
        {"stay": obj, "stars": [1, 2, 3, 4, 5], "all_fields": all_fields},
    )

def stay_add(request):
    if request.method == "POST":
        form = StayForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            # Normalize state and auto-geocode if missing coords
            st = getattr(obj, "state", None)
            if isinstance(st, str):
                obj.state = st.strip().upper()
            if (getattr(obj, "latitude", None) is None) or (getattr(obj, "longitude", None) is None):
                try:
                    from .utils import build_query_from_stay, geocode_address
                    q = build_query_from_stay(obj)
                    if q:
                        coords = geocode_address(q)
                        if coords:
                            obj.latitude, obj.longitude = coords
                except Exception:
                    pass
            obj.save()
            return redirect("stays:detail", pk=obj.pk)
    else:
        form = StayForm()
    return render(request, "stays/stay_form.html", {"form": form})

def stay_edit(request, pk):
    obj = get_object_or_404(Stay, pk=pk)
    if request.method == "POST":
        form = StayForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            obj = form.save(commit=False)
            st = getattr(obj, "state", None)
            if isinstance(st, str):
                obj.state = st.strip().upper()
            if (getattr(obj, "latitude", None) is None) or (getattr(obj, "longitude", None) is None):
                try:
                    from .utils import build_query_from_stay, geocode_address
                    q = build_query_from_stay(obj)
                    if q:
                        coords = geocode_address(q)
                        if coords:
                            obj.latitude, obj.longitude = coords
                except Exception:
                    pass
            obj.save()
            return redirect("stays:detail", pk=obj.pk)
    else:
        form = StayForm(instance=obj)
    return render(request, "stays/stay_form.html", {"form": form, "stay": obj})

def stay_geocode(request, pk):
    obj = get_object_or_404(Stay, pk=pk)
    if request.method != 'POST':
        return HttpResponseBadRequest('POST required')
    try:
        from .utils import geocode_from_stay
        coords = geocode_from_stay(obj)
        if coords:
            obj.latitude, obj.longitude = coords
            obj.save(update_fields=['latitude', 'longitude'])
            messages.success(request, f"Geocoded to {coords[0]:.6f}, {coords[1]:.6f}.")
        else:
            messages.warning(request, "Could not geocode this stay.")
    except Exception as e:
        messages.error(request, f"Geocode error: {e}")
    return redirect('stays:edit', pk=obj.pk)

# Destructive: delete a Stay (POST only)
def stay_delete(request, pk):
    obj = get_object_or_404(Stay, pk=pk)
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")
    obj.delete()
    messages.success(request, "Stay deleted.")
    return redirect("stays:list")

# rating is edited only on add/edit forms; no list-page endpoint

# --- Charts (basic, graceful if fields missing) ---
def stay_charts(request):
    """Three charts: by State, by Year (if date field exists), Rating distribution (if rating exists)."""
    field_names = {getattr(f, "attname", None) or getattr(f, "name", None) for f in Stay._meta.get_fields()}

    # 1) by State
    state_counts = Stay.objects.values_list("state").exclude(state__isnull=True).exclude(state__exact="")
    state_map = {}
    for s, in state_counts:
        state_map[s] = state_map.get(s, 0) + 1
    states = sorted(state_map.keys())
    states_series = [state_map[s] for s in states]

    # 2) by Year (best-effort)
    years_labels, years_series = [], []
    date_field = None
    for cand in ("date", "start_date", "arrival_date", "created", "created_at", "updated", "updated_at"):
        if cand in field_names:
            date_field = cand
            break
    if date_field:
        from django.db.models.functions import ExtractYear
        qs_years = (Stay.objects.exclude(**{f"{date_field}__isnull": True})
                    .annotate(y=ExtractYear(date_field))
                    .values_list("y"))
        ym = {}
        for y, in qs_years:
            if y is None:
                continue
            ym[y] = ym.get(y, 0) + 1
        for y in sorted(ym.keys()):
            years_labels.append(str(y))
            years_series.append(ym[y])

    # 3) Rating distribution (fully guarded for missing column)
    rating_labels, rating_series = [], []
    if "rating" in field_names:
        rm = {}
        try:
            vals = list(Stay.objects.exclude(rating__isnull=True).values_list("rating", flat=True))
            for v in vals:
                try:
                    iv = int(v)
                except Exception:
                    continue
                if 1 <= iv <= 5:
                    rm[iv] = rm.get(iv, 0) + 1
        except Exception:
            rm = {}
        for r in sorted(rm.keys()):
            rating_labels.append(str(r))
            rating_series.append(rm[r])

    ctx = {
        "states": json.dumps(states),
        "states_series": json.dumps(states_series),
        "years_labels": json.dumps(years_labels),
        "years_series": json.dumps(years_series),
        "rating_labels": json.dumps(rating_labels),
        "rating_series": json.dumps(rating_series),
        "has_years": bool(years_labels),
        "has_rating": bool(rating_labels),
    }
    return render(request, "stays/charts.html", ctx)

# --- Import/Export CSV ---
def import_stays_csv(request):
    """Upload a CSV and create Stay rows. Accepts flexible headers.

    Supported headers (case-insensitive, trimmed):
    - Park, City, State, City/St, Check in, Leave, #Nts, Rate/nt, Total, Fees, Paid?, Site, Notes
    Unknown headers are ignored. Dates accept YYYY-MM-DD, MM/DD/YYYY.
    """
    if request.method == "GET":
        # Allow rendering as an options page as well
        return render(request, "stays/import.html", {
            "default_delimiter": request.GET.get("delimiter", "auto"),
            "default_dry": request.GET.get("dry_run") in {"1", "true", "yes"},
        })

    # POST
    f = request.FILES.get("file")
    if not f:
        return render(request, "stays/import.html", {"error": "Please choose a CSV file."})

    # Determine type and build rows list of dicts
    created = 0
    updated_count = 0
    skipped = 0
    auto_geocode = (request.POST.get("autogeocode") or request.GET.get("autogeocode")) in {"1", "true", "yes"}
    dry_run = (request.POST.get("dry_run") or request.GET.get("dry_run")) in {"1", "true", "yes"}

    filename = (getattr(f, "name", "") or "").lower()
    sheet_name = (request.POST.get("sheet") or request.GET.get("sheet") or "").strip()
    rows = []
    header_keys = {}

    if filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        try:
            import openpyxl  # type: ignore
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True, read_only=True)
            ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
            header = None
            header_row_index = None
            row_index = 0
            for r in ws.iter_rows(values_only=True):
                row_index += 1
                if r and any(v is not None and str(v).strip() != '' for v in r):
                    header = [str(v).strip() if v is not None else '' for v in r]
                    header_row_index = row_index
                    break
            if not header:
                return render(request, "stays/import.html", {"error": "Could not find a header row in the selected sheet."})
            # Build row dicts
            start_row = (header_row_index or 1) + 1
            for r in ws.iter_rows(min_row=start_row, values_only=True):
                if r is None:
                    continue
                row = {}
                for i, key in enumerate(header):
                    if not key:
                        continue
                    val = r[i] if i < len(r) else None
                    row[key] = '' if val is None else str(val)
                if any((str(v).strip() for v in row.values())):
                    rows.append(row)
            header_keys = { (h or '').strip().lower(): h for h in header }
        except Exception as e:
            return render(request, "stays/import.html", {"error": f"Failed to read Excel file: {e}"})
    else:
        # CSV path
        raw = f.read()
        try:
            text = raw.decode("utf-8")
        except Exception:
            text = raw.decode("latin-1", errors="ignore")

        # Determine delimiter
        delim_choice = (request.POST.get("delimiter") or request.GET.get("delimiter") or "auto").lower()
        delimiter = ","
        if delim_choice in {"comma", ","}:
            delimiter = ","
        elif delim_choice in {"semicolon", ";"}:
            delimiter = ";"
        elif delim_choice in {"tab", "\t"}:
            delimiter = "\t"
        else:
            # auto-sniff
            try:
                sniffer = csv.Sniffer()
                sample = text[:2048]
                dialect = sniffer.sniff(sample, delimiters=[",", ";", "\t"])
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ","

        reader = csv.DictReader(text.splitlines(), delimiter=delimiter)
        rows = list(reader)
        def norm(h):
            return (h or "").strip().lower()
        header_keys = {norm(h): h for h in (reader.fieldnames or [])}

    def norm(h):
        return (h or "").strip().lower()
    def norm2(h):
        # looser normalization for fuzzy header matching
        return ''.join(ch for ch in norm(h) if ch.isalnum())

    def get(row, *cands):
        # try exact header map first
        for c in cands:
            key = header_keys.get(norm(c))
            if key and key in row:
                val = row.get(key)
                if val is None:
                    continue
                s = str(val).strip()
                if s != "":
                    return s
        # fuzzy: match by normalized containment (e.g., "park name" -> "park")
        row_keys = list(row.keys())
        for c in cands:
            nc = norm2(c)
            if not nc:
                continue
            for rk in row_keys:
                if not rk:
                    continue
                nr = norm2(rk)
                if not nr:
                    continue
                if nc in nr or nr in nc:
                    val = row.get(rk)
                    if val is None:
                        continue
                    s = str(val).strip()
                    if s != "":
                        return s
        return ""

    def parse_date(val):
        if not val:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(val, fmt).date()
            except Exception:
                pass
        return None

    def parse_money(val):
        if not val:
            return None
        s = val.replace(",", "").replace("$", "").strip()
        try:
            from decimal import Decimal
            return Decimal(s)
        except Exception:
            return None

    def parse_coord(val):
        if not val:
            return None
        s = str(val).strip()
        try:
            from decimal import Decimal
            return Decimal(s)
        except Exception:
            try:
                return Decimal(str(float(s)))
            except Exception:
                return None

    def parse_bool(val):
        s = (val or "").strip().lower()
        return s in {"y", "yes", "true", "1"}

    def parse_int_1_5(val):
        try:
            i = int((val or "").strip())
            if 1 <= i <= 5:
                return i
        except Exception:
            pass
        return None

    # Prepare skipped report writer
    skipped_rows = []
    def record_skip(reason: str, raw_row: dict):
        nonlocal skipped_rows
        row_copy = dict(raw_row)
        row_copy["__reason__"] = reason
        skipped_rows.append(row_copy)

    dedupe_mode = (request.POST.get("dedupe") or request.GET.get("dedupe") or "skip").lower()

    for row in rows:
        # Support combined City/St column like "Austin, TX" or "Austin/TX" or "Austin TX"
        city_st = get(row, "City/St", "City/State")
        city = get(row, "City")
        state = get(row, "State")
        if (not city or not state) and city_st:
            raw = city_st.strip()
            # Normalize common separators to comma
            for sep in ["/", "|", " - ", "-"]:
                raw = raw.replace(sep, ",")
            parts = [p.strip() for p in raw.split(",") if p.strip()]
            if len(parts) >= 2:
                if not city:
                    city = parts[0]
                if not state:
                    token = parts[1]
                    # If looks like full state name, take first 2 letters; otherwise trust token
                    state = token[:2]
            elif len(parts) == 1 and " " in parts[0]:
                # Fall back: split on last space -> "CityName ST"
                base = parts[0]
                cparts = base.rsplit(" ", 1)
                if len(cparts) == 2:
                    if not city:
                        city = cparts[0].strip()
                    if not state:
                        state = cparts[1].strip()[:2]

        # If city field itself has trailing ", ST", split and adopt state if missing
        if city:
            if "," in city:
                parts_city = [p.strip() for p in city.split(",")]
                if len(parts_city) >= 2:
                    # If trailing token is 2-letter, adopt as state when missing
                    if len(parts_city[-1]) == 2 and not state:
                        state = parts_city[-1]
                    # Always reduce city to the portion before the first comma
                    city = parts_city[0]
            else:
                # Handle "City ST" without comma
                tokens = city.strip().split()
                if len(tokens) >= 2 and len(tokens[-1]) == 2:
                    if not state:
                        state = tokens[-1]
                    city = " ".join(tokens[:-1]).strip()

        obj = Stay(
            park=get(row, "Park", "Campground", "Name", "Park Name", "Camp Name", "Campground Name", "Location", "Place"),
            city=city,
            state=(state or "")[:2].upper(),
            check_in=parse_date(get(row, "Check in", "Check-in", "Arrival", "Start")),
            leave_date=parse_date(get(row, "Leave", "Check out", "Departure", "End")),
            price_night=parse_money(get(row, "Rate/nt", "Rate", "Price/night")),
            total=parse_money(get(row, "Total")),
            fees=parse_money(get(row, "Fees", "Taxes/fees")),
            paid=parse_bool(get(row, "Paid?", "Paid")),
            rating=parse_int_1_5(get(row, "Rating")),
            site=get(row, "Site"),
            notes=get(row, "Notes"),
        )

        # Skip completely empty location/date rows: no City, no State, no Check in, no Leave
        empty_loc_dates = (not (obj.city or "").strip()) and (not (obj.state or "").strip()) \
                          and (obj.check_in is None) and (obj.leave_date is None)
        if empty_loc_dates:
            skipped += 1
            record_skip("empty city/state/dates", row)
            continue
        # Optional latitude/longitude columns
        lat = parse_coord(get(row, "Latitude", "Lat"))
        lng = parse_coord(get(row, "Longitude", "Long", "Lng"))
        if lat is not None:
            obj.latitude = lat
        if lng is not None:
            obj.longitude = lng
        if not dry_run:
            # Skip if a duplicate exists (same park/city/state/check_in/leave_date)
            try:
                norm_park = (obj.park or "").strip()
                norm_city = (obj.city or "").strip()
                norm_state = (obj.state or "").strip().upper()[:2]
                strong_key = bool(norm_park and norm_city and norm_state and obj.check_in and obj.leave_date)
                if strong_key:
                    existing_qs = Stay.objects.filter(
                        park__iexact=norm_park,
                        city__iexact=norm_city,
                        state=norm_state,
                        check_in=obj.check_in,
                        leave_date=obj.leave_date,
                    )
                    if existing_qs.exists():
                        if dedupe_mode == "skip":
                            skipped += 1
                            record_skip("duplicate", row)
                            continue
                        elif dedupe_mode == "update":
                            # Update first matching row
                            target = existing_qs.first()
                            fields = [
                                "price_night","total","fees","paid","rating","site","notes",
                                "latitude","longitude","park","city","state"
                            ]
                            for f in fields:
                                setattr(target, f, getattr(obj, f))
                            target.save()
                            updated_count += 1
                            continue
                        # else: 'none' falls through to create a new row
            except Exception:
                pass
            if auto_geocode and (obj.latitude is None or obj.longitude is None):
                try:
                    from .utils import geocode_from_stay
                    coords = geocode_from_stay(obj)
                    if coords:
                        obj.latitude, obj.longitude = coords
                except Exception:
                    pass
            obj.save()
        created += 1

    # If we have skipped rows, write a CSV into MEDIA_ROOT/exports and expose URL
    skipped_url = None
    if skipped_rows and not dry_run:
        try:
            base = getattr(settings, "EXPORTS_DIR", None)
            if not base:
                media = getattr(settings, "MEDIA_ROOT", None)
                base = Path(media) / "exports" if media else Path(settings.BASE_DIR) / "exports"
            base = Path(base)
            base.mkdir(parents=True, exist_ok=True)
            from io import StringIO
            import csv as _csv
            buf = StringIO()
            # Determine headers from first row
            keys = list(skipped_rows[0].keys())
            writer = _csv.DictWriter(buf, fieldnames=keys)
            writer.writeheader()
            for r in skipped_rows:
                writer.writerow(r)
            fname = f"skipped_import_{datetime.now().strftime('%Y%m%d-%H%M%S')}.csv"
            target = base / fname
            target.write_text(buf.getvalue(), encoding="utf-8")
            # Build URL under MEDIA_URL if inside MEDIA_ROOT, else show absolute path
            media = getattr(settings, "MEDIA_ROOT", None)
            media_url = getattr(settings, "MEDIA_URL", "/media/")
            if media and str(target).startswith(str(Path(media))):
                rel = str(Path(target).relative_to(Path(media)))
                skipped_url = str(Path(media_url) / Path(rel))
            else:
                skipped_url = None
        except Exception:
            skipped_url = None

    return render(request, "stays/import_result.html", {
        "created": created,
        "updated": updated_count,
        "skipped": skipped,
        "skipped_url": skipped_url,
        "dry_run": dry_run,
    })


def export_stays_csv(request):
    """Stream all stays as CSV. Optionally save to server disk.

    Query params:
    - save=1           -> also save to disk
    - filename=name    -> target file name (default: stays_export.csv)
    - subdir=path      -> optional subdirectory under EXPORTS base

    Base directory: settings.EXPORTS_DIR or MEDIA_ROOT/exports.
    Paths are sanitized and constrained to the base directory.
    """
    headers = [
        "Park", "City", "State", "Check in", "Leave", "#Nts",
        "Rate/nt", "Total", "Fees", "Paid?", "Rating", "Site", "Notes",
    ]

    # Always prepare CSV in-memory first so we can both save and/or stream.
    from io import StringIO
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)

    for s in Stay.objects.all().order_by("-check_in", "-id"):
        writer.writerow([
            smart_str(s.park or ""),
            smart_str(s.city or ""),
            smart_str((s.state or "").upper()[:2]),
            s.check_in.isoformat() if s.check_in else "",
            s.leave_date.isoformat() if s.leave_date else "",
            s.nights or 0,
            f"{s.price_night:.2f}" if s.price_night is not None else "",
            f"{s.total:.2f}" if s.total is not None else "",
            f"{s.fees:.2f}" if s.fees is not None else "",
            "Yes" if s.paid else "No",
            str(getattr(s, 'rating', '') or ""),
            smart_str(s.site or ""),
            smart_str(s.notes or ""),
        ])
    csv_text = buffer.getvalue()

    # If requested, save to disk within a safe base directory.
    saved_path = None
    if request.GET.get("save") in {"1", "true", "yes"}:
        base = getattr(settings, "EXPORTS_DIR", None)
        if not base:
            # default to MEDIA_ROOT/exports or BASE_DIR/exports
            media = getattr(settings, "MEDIA_ROOT", None)
            base = Path(media) / "exports" if media else Path(settings.BASE_DIR) / "exports"
        base = Path(base)
        base.mkdir(parents=True, exist_ok=True)

        # Sanitize filename and subdir
        import re
        filename = request.GET.get("filename") or "stays_export.csv"
        filename = re.sub(r"[^A-Za-z0-9._-]", "_", filename)
        subdir = request.GET.get("subdir", "").strip()
        subdir = re.sub(r"[^A-Za-z0-9._/-]", "_", subdir).strip("/")

        target = base / subdir / filename if subdir else base / filename
        # Resolve to avoid directory traversal
        target = target.resolve()
        if not str(target).startswith(str(base.resolve())):
            return HttpResponseBadRequest("Invalid path.")
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(csv_text, encoding="utf-8")
        saved_path = str(target)

    # Stream response to client
    download_name = request.GET.get("filename") or "stays_export.csv"
    response = HttpResponse(csv_text, content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{download_name}"'
    if saved_path:
        response["X-Saved-To"] = saved_path
    return response


def export_stays_options(request):
    """Render a small form to choose export filename/subfolder and saving behavior."""
    # Provide simple defaults
    default_name = request.GET.get("filename") or "stays_export.csv"
    default_subdir = request.GET.get("subdir") or ""
    default_save = request.GET.get("save") in {"1", "true", "yes"}
    # Compute display base path
    base = getattr(settings, "EXPORTS_DIR", None)
    if not base:
        media = getattr(settings, "MEDIA_ROOT", None)
        base = Path(media) / "exports" if media else Path(settings.BASE_DIR) / "exports"
    base = str(Path(base))
    return render(request, "stays/export.html", {
        "default_name": default_name,
        "default_subdir": default_subdir,
        "default_save": default_save,
        "exports_base": base,
    })


def import_stays_options(request):
    """Render import page with options (delimiter, dry run)."""
    return render(request, "stays/import.html", {
        "default_delimiter": request.GET.get("delimiter", "auto"),
        "default_dry": request.GET.get("dry_run") in {"1", "true", "yes"},
    })


def import_probe_sheets(request):
    """Return Excel sheet names for a posted .xlsx file.

    POST: multipart with 'file'
    Response: { sheets: [..] } or { error: '...' }
    """
    if request.method != 'POST':
        return JsonResponse({"error": "POST required"}, status=400)
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({"error": "Missing file"}, status=400)
    name = (getattr(f, 'name', '') or '').lower()
    if not name.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return JsonResponse({"error": "Not an Excel file"}, status=400)
    try:
        import openpyxl  # type: ignore
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=True)
        return JsonResponse({"sheets": wb.sheetnames})
    except Exception as e:
        return JsonResponse({"error": f"Failed to read Excel: {e}"}, status=400)
